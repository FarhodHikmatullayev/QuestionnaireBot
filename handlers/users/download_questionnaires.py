import os

from aiogram import types
from aiogram.dispatcher import FSMContext

from data.config import ADMINS
from keyboards.default.all_branches import all_branches_default_keyboard
from keyboards.default.menu_keyboards import back_to_menu
from loader import dp, db
from states.download_state import DownloadQuestionnairesState

import tempfile
from openpyxl.styles import Alignment
import openpyxl


async def download_questionnaires(branch_id):
    questionnaires = await db.select_questionnaires(branch_id=branch_id)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = "Filial"
    worksheet['C1'] = "Bo'lim (Lavozim)"
    worksheet['D1'] = 'Ish muhiti va madaniyati'
    worksheet['E1'] = "Rivojlanish va o'sish imkoniyatlari"
    worksheet['F1'] = 'Ish haqi va mukofotlar'
    worksheet['G1'] = 'Rahbariyat bilan munosabat'
    worksheet[
        'H1'] = "Ish muvozanati va farovonligi (Ish va hayot muvozanatiga e'tibor, stress darajasi, dam olish va tanaffuslar, ish yuklamasi)"

    for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']:
        worksheet[cell].alignment = Alignment(horizontal='center')

    worksheet.cell(row=1, column=1, value='T/r')
    worksheet.cell(row=1, column=2, value='Filial')
    worksheet.cell(row=1, column=3, value="Bo'lim (Lavozim)")
    worksheet.cell(row=1, column=4, value='Ish muhiti va madaniyati')
    worksheet.cell(row=1, column=5, value="Rivojlanish va o'sish imkoniyatlari")
    worksheet.cell(row=1, column=6, value='Ish haqi va mukofotlar')
    worksheet.cell(row=1, column=7, value='Rahbariyat bilan munosabat')
    worksheet.cell(row=1, column=8,
                   value="Ish muvozanati va farovonligi (Ish va hayot muvozanatiga e'tibor, stress darajasi, dam olish va tanaffuslar, ish yuklamasi)")

    branch = await db.select_branch(branch_id=branch_id)
    branch_name = branch['name']

    tr = 0
    for questionnaire in questionnaires:
        tr += 1
        rank_id = questionnaire['rank_id']
        rank = await db.select_rank(rank_id=rank_id)
        rank_name = rank['name']
        ish_muhiti_va_madaniyati = questionnaire['ish_muhiti_va_madaniyati']
        rivojlanish_va_osish_imkoniyatlari = questionnaire['rivojlanish_va_osish_imkoniyatlari']
        ish_haqi_va_mukofotlar = questionnaire['ish_haqi_va_mukofotlar']
        rahbariyat_bilan_munosabat = questionnaire['rahbariyat_bilan_munosabat']
        ish_muvozanati_va_farovonligi = questionnaire['ish_muvozanati_va_farovonligi']

        worksheet.cell(row=tr + 1, column=1, value=tr).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=2, value=branch_name).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=3, value=rank_name).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=4, value=ish_muhiti_va_madaniyati).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=5, value=rivojlanish_va_osish_imkoniyatlari).alignment = Alignment(
            horizontal='center')
        worksheet.cell(row=tr + 1, column=6, value=ish_haqi_va_mukofotlar).alignment = Alignment(
            horizontal='center')
        worksheet.cell(row=tr + 1, column=7, value=rahbariyat_bilan_munosabat).alignment = Alignment(
            horizontal='center')
        worksheet.cell(row=tr + 1, column=8, value=ish_muvozanati_va_farovonligi).alignment = Alignment(
            horizontal='center')

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'sorovnoma.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(text="📥 Javoblarni yuklab olish", state="*", user_id=ADMINS)
async def get_branch(message: types.Message, state: FSMContext):
    try:
        await state.finish()
    except:
        pass
    markup = await all_branches_default_keyboard()
    await message.answer(text="Qaysi filial uchun javoblarni yuklab olmoqchisiz?\n"
                              "Filialni tanlang 👇", reply_markup=markup)
    await DownloadQuestionnairesState.branch_id.set()


@dp.message_handler(text="📥 Javoblarni yuklab olish", state="*")
async def get_branch(message: types.Message, state: FSMContext):
    try:
        await state.finish()
    except:
        pass
    await message.reply(text="Sizga bu funksiyadan foydalanishda ruxsat mavjud emas", reply_markup=back_to_menu)


@dp.message_handler(state=DownloadQuestionnairesState.branch_id)
async def get_branch_id(message: types.Message, state: FSMContext):
    branch_name = message.text
    branches = await db.select_branches(name=branch_name)
    if not branches:
        await message.reply(text="❌ Bunday filial topilmadi. 😔", reply_markup=back_to_menu)
        return
    branch = branches[0]
    branch_id = branch['id']
    await state.update_data(branch_id=branch_id)

    temp_dir = await download_questionnaires(branch_id=branch_id)

    with open(os.path.join(temp_dir, 'sorovnoma.xlsx'), 'rb') as file:
        await message.answer_document(document=file)

    os.remove(os.path.join(temp_dir, 'sorovnoma.xlsx'))
